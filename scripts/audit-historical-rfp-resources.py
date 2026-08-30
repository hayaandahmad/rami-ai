#!/usr/bin/env python3
"""Read-only audit of historical RFP Question Bank Excel workbooks.
Does not modify source files. Does not touch ProjectFacts / PostgreSQL."""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

CANONICAL_QIDS = {
    "0.1", "0.2", "0.3", "0.4", "0.5", "0.6", "0.7",
    "1.1", "1.2", "1.3", "1.4", "1.5",
    "2.1", "2.2", "2.3",
    "3.1", "3.2", "3.3", "3.4", "3.5",
    "4.1", "4.2", "4.3", "4.4", "4.5",
    "5.1", "5.2", "5.3", "5.4", "5.5",
    "6.1", "6.2", "6.3", "6.4", "6.5", "6.6",
    "7.1", "7.2", "7.3",
    "8.1", "8.2", "8.3", "8.4",
    "9.1", "9.2", "9.3", "9.4",
    "10.1", "10.2", "10.3", "10.4",
    "11.1", "11.2", "11.3",
    "12.1", "12.2", "12.3", "12.4", "12.5", "12.6", "12.7", "12.8",
}

Q_TO_FIELDS: dict[str, list[str]] = {
    "0.1": ["documentType"], "0.2": ["documentTitle"], "0.3": ["beneficiaryEntity"],
    "0.4": ["tenderNumber"], "0.5": ["proposalDeadline"], "0.6": [], "0.7": ["referenceTemplateId"],
    "1.1": ["currentSituation"], "1.2": ["painPoints"], "1.3": ["businessNeedRationale"],
    "1.4": ["businessObjectives"], "1.5": ["previousPhases"],
    "2.1": ["engagementType"], "2.2": ["engagementPhases"], "2.3": ["engagementDuration"],
    "3.1": ["beneficiaryEntity"], "3.2": ["users"], "3.3": ["stakeholderRoles"],
    "3.4": ["approvers", "uatOwners"], "3.5": ["postGoLiveOwner"],
    "4.1": ["inScope"], "4.2": ["outOfScope"], "4.3": ["bidderResponsibilities"],
    "4.4": ["entityResponsibilities"], "4.5": ["assumptionsDependenciesConstraints"],
    "5.1": ["functionalModules"], "5.2": ["keyWorkflows"], "5.3": ["reportingNeeds"],
    "5.4": ["caseManagementNeeds"], "5.5": ["aiFeatures"],
    "6.1": ["hostingModel"], "6.2": ["integrations"], "6.3": ["securityRequirements"],
    "6.4": ["performanceAvailabilityTargets"], "6.5": ["dataMigrationNeeds"], "6.6": [],
    "7.1": ["deliverableItems"], "7.2": ["deliverableFormats"], "7.3": ["deliverableApprovers"],
    "8.1": ["engagementPhases", "engagementDuration"], "8.2": ["uatRounds"],
    "8.3": ["acceptanceCriteria"], "8.4": ["rollbackPlanNeeded"],
    "9.1": ["supportPeriodAndHours"], "9.2": ["slaTiers"], "9.3": ["supportOperatingModel"],
    "9.4": ["supportPenalties"],
    "10.1": ["evaluationWeights"], "10.2": ["evaluationRules"],
    "10.3": ["pricingModelAndCostBreakdown"], "10.4": ["optionalItemsAndTaxes"],
    "11.1": ["legalTerms"], "11.2": ["jvSubcontractingRules"], "11.3": ["requiredAnnexes"],
    "12.1": ["riskNotes"], "12.2": ["riskNotes"], "12.3": ["riskNotes"], "12.4": [],
    "12.5": ["optionalItemsAndTaxes"], "12.6": [], "12.7": [], "12.8": ["requiredAnnexes"],
}

FIELD_SECTIONS: dict[str, list[str]] = {
    "documentType": ["coverPage"], "documentTitle": ["coverPage"],
    "beneficiaryEntity": ["coverPage", "introduction"],
    "tenderNumber": ["coverPage"], "proposalDeadline": ["coverPage", "administrativeProcedures"],
    "referenceTemplateId": ["coverPage"],
    "currentSituation": ["background"], "painPoints": ["background"],
    "businessNeedRationale": ["background"], "businessObjectives": ["background"],
    "previousPhases": ["background"], "riskNotes": ["background"],
    "engagementType": ["engagementDefinition"],
    "engagementPhases": ["engagementDefinition", "implementationRequirements"],
    "engagementDuration": ["engagementDefinition", "implementationRequirements"],
    "users": ["introduction"], "stakeholderRoles": ["introduction", "manpower"],
    "approvers": ["introduction"], "uatOwners": ["introduction", "acceptanceCriteria"],
    "postGoLiveOwner": ["introduction"],
    "inScope": ["scopeOfWork"], "outOfScope": ["scopeOfWork"],
    "bidderResponsibilities": ["scopeOfWork"], "entityResponsibilities": ["scopeOfWork"],
    "assumptionsDependenciesConstraints": ["scopeOfWork"],
    "functionalModules": ["functionalRequirements"], "keyWorkflows": ["functionalRequirements"],
    "reportingNeeds": ["functionalRequirements"], "caseManagementNeeds": ["functionalRequirements"],
    "aiFeatures": ["functionalRequirements"],
    "hostingModel": ["technicalRequirements"], "integrations": ["technicalRequirements"],
    "securityRequirements": ["technicalRequirements"],
    "performanceAvailabilityTargets": ["technicalRequirements"],
    "dataMigrationNeeds": ["technicalRequirements"],
    "deliverableItems": ["deliverables"], "deliverableFormats": ["deliverables"],
    "deliverableApprovers": ["deliverables"],
    "uatRounds": ["acceptanceCriteria"], "acceptanceCriteria": ["acceptanceCriteria"],
    "rollbackPlanNeeded": ["acceptanceCriteria"],
    "supportPeriodAndHours": ["supportMaintenance"], "slaTiers": ["supportMaintenance"],
    "supportOperatingModel": ["supportMaintenance"],
    "supportPenalties": ["supportMaintenance", "legalContractualTerms"],
    "evaluationWeights": ["evaluationCriteria"], "evaluationRules": ["evaluationCriteria"],
    "pricingModelAndCostBreakdown": ["financialProposal"],
    "optionalItemsAndTaxes": ["financialProposal"],
    "legalTerms": ["legalContractualTerms"], "jvSubcontractingRules": ["legalContractualTerms"],
    "requiredAnnexes": ["annexes"],
}

EMPTY = {"", "n/a", "na", "n.a.", "none", "not found", "not available", "-", "—", "unknown"}
STRONG_STATUS = {"answered", "found", "extracted", "confirmed", "present"}
PARTIAL_STATUS = {"partially stated", "partial", "tbc", "to be confirmed", "deferred", "not applicable"}

SECRET_PATTERNS = [
    re.compile(r"(?i)(api[_-]?key|password|secret|token|bearer)\s*[:=]\s*\S+"),
    re.compile(r"(?i)sk-[a-zA-Z0-9]{20,}"),
    re.compile(r"(?i)-----BEGIN (RSA |OPENSSH )?PRIVATE KEY-----"),
]

GAP_KEYWORDS = {
    "namedRoles": [r"named role", r"key personnel", r"project manager", r"pmp", r"staff roles", r"cv\b"],
    "clarificationContact": [r"clarification", r"enquir", r"contact person", r"questions to"],
    "submissionChannel": [r"joneps", r"e-?procurement", r"submission portal", r"submit proposal"],
    "governanceCadence": [r"steering committee", r"governance", r"weekly meeting", r"progress report", r"pmo"],
    "knowledgeTransfer": [r"knowledge transfer", r"training of trainers", r"handover"],
    "implementationStages": [r"phase\s*\d", r"stage\s*\d", r"milestone", r"workstream"],
    "procurementAdmin": [r"bid bond", r"performance bond", r"eligibility", r"qualification", r"award"],
    "manpower": [r"manpower", r"fte", r"staffing", r"on-?site", r"resident engineer", r"minimum staff"],
    "callOffSow": [r"call-?off", r"work order", r"sow\b", r"assignment"],
    "bidderCountAward": [r"top\s*\d", r"number of (bidders|suppliers|winners)", r"award model"],
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def is_empty(s: str) -> bool:
    return s.strip().lower() in EMPTY or not s.strip()


def find_header_row(rows: list) -> tuple[int, list[str]] | tuple[None, None]:
    for i, r in enumerate(rows):
        vals = [cell_str(c) for c in (r or [])]
        norms = [v.lower() for v in vals]
        if any(n == "question id" for n in norms) and any("question" in n for n in norms):
            return i, vals
    return None, None


def map_columns(headers: list[str]) -> dict[str, int]:
    norms = [re.sub(r"\s+", " ", h.strip().lower()) for h in headers]
    out: dict[str, int] = {}
    for i, n in enumerate(norms):
        if n == "question id" or n == "qid":
            out["question_id"] = i
        elif n in {"exact rami question", "question", "question text", "suggested question"}:
            out["question_text"] = i
        elif "answer" in n:
            out["answer"] = i
        elif n == "status":
            out["status"] = i
        elif n.startswith("source"):
            out["source"] = i
        elif n == "section":
            out["section"] = i
        elif n in {"notes", "note", "comments"}:
            out["notes"] = i
    return out


def parse_sheet(rows: list, sheet_name: str) -> dict:
    hdr_i, headers = find_header_row(rows)
    if hdr_i is None:
        return {
            "name": sheet_name,
            "kind": "unstructured_or_title_only",
            "rowCount": 0,
            "headers": [],
            "rows": [],
        }
    cols = map_columns(headers)
    data = []
    secret_hits = 0
    for r in rows[hdr_i + 1 :]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        vals = [cell_str(c) for c in r]
        joined = " | ".join(vals)
        for pat in SECRET_PATTERNS:
            if pat.search(joined):
                secret_hits += 1
        qid = vals[cols["question_id"]] if "question_id" in cols else ""
        qid = re.sub(r"^q\s*", "", qid.strip(), flags=re.I)
        qid = re.sub(r"\s+", "", qid)
        ans = vals[cols["answer"]] if "answer" in cols else ""
        status = vals[cols["status"]] if "status" in cols else ""
        source = vals[cols["source"]] if "source" in cols else ""
        qtext = vals[cols["question_text"]] if "question_text" in cols else ""
        section = vals[cols["section"]] if "section" in cols else ""
        if not qid and not qtext:
            continue
        data.append({
            "questionId": qid,
            "questionText": qtext,
            "answer": ans,
            "status": status,
            "source": source,
            "section": section,
            "answered": not is_empty(ans),
        })
    kind = "qa"
    if "suggested" in sheet_name.lower():
        kind = "suggested_additions"
    elif "adaptive" in sheet_name.lower():
        kind = "adaptive_depth"
    elif "architecture" in sheet_name.lower():
        kind = "architecture_notes"
    return {
        "name": sheet_name,
        "kind": kind,
        "headerRowIndex": hdr_i + 1,
        "headers": headers,
        "columnMap": cols,
        "rowCount": len(data),
        "secretPatternHits": secret_hits,
        "rows": data,
    }


def classify_field_support(qa_rows: list[dict]) -> tuple[set[str], set[str], set[str]]:
    strong: set[str] = set()
    partial: set[str] = set()
    unsupported_candidates: set[str] = set()
    for fs in Q_TO_FIELDS.values():
        unsupported_candidates.update(fs)

    for r in qa_rows:
        qid = r["questionId"]
        fields = Q_TO_FIELDS.get(qid, [])
        if not fields:
            continue
        st = r["status"].strip().lower()
        ans = r["answer"]
        if not r["answered"]:
            continue
        if st in STRONG_STATUS or (len(ans) >= 60 and st not in PARTIAL_STATUS and "tbc" not in ans.lower()[:40]):
            if st in PARTIAL_STATUS or ans.lower().startswith("[to be confirmed]") or st == "tbc":
                partial.update(fields)
            else:
                strong.update(fields)
        elif st in PARTIAL_STATUS or "tbc" in st or ans.lower().startswith("[to be confirmed]"):
            partial.update(fields)
        else:
            partial.update(fields)
    partial -= strong
    unsupported = unsupported_candidates - strong - partial
    return strong, partial, unsupported


def audit_workbook(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        sheets.append(parse_sheet(rows, name))
    wb.close()

    qa = next((s for s in sheets if s["kind"] == "qa"), None)
    suggested = next((s for s in sheets if s["kind"] == "suggested_additions"), None)
    qa_rows = qa["rows"] if qa else []

    qids = [r["questionId"] for r in qa_rows if r["questionId"]]
    matched = sorted(set(q for q in qids if q in CANONICAL_QIDS))
    extra = sorted(set(q for q in qids if q and q not in CANONICAL_QIDS))
    missing = sorted(CANONICAL_QIDS - set(matched))

    answered = []
    for r in qa_rows:
        if r["questionId"] not in CANONICAL_QIDS:
            continue
        st = r["status"].strip().lower()
        if r["answered"] and st not in {"not applicable"}:
            # count NA separately
            answered.append(r["questionId"])
        elif r["answered"]:
            answered.append(r["questionId"])
    answered_qids = sorted(set(answered))

    status_counts = Counter(r["status"] for r in qa_rows if r["status"])
    with_source = sum(1 for r in qa_rows if r["source"] and not is_empty(r["source"]))
    with_status = sum(1 for r in qa_rows if r["status"] and not is_empty(r["status"]))

    strong, partial, unsupported = classify_field_support(qa_rows)
    section_hits: Counter = Counter()
    for f in strong | partial:
        for s in FIELD_SECTIONS.get(f, []):
            section_hits[s] += 1

    blob = "\n".join(
        f"{r['answer']}\n{r['questionText']}\n{r['status']}\n{r['source']}" for r in qa_rows
    )
    if suggested:
        blob += "\n" + "\n".join(
            f"{r['answer']}\n{r['questionText']}" for r in suggested["rows"]
        )
    gap_evidence = {}
    for gap, pats in GAP_KEYWORDS.items():
        hits = sum(1 for p in pats if re.search(p, blob, re.I))
        gap_evidence[gap] = hits

    suggested_ids = []
    if suggested:
        suggested_ids = sorted({r["questionId"] for r in suggested["rows"] if r["questionId"]})

    return {
        "file": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256(path),
        "sheetNames": [s["name"] for s in sheets],
        "sheetsSummary": [
            {
                "name": s["name"],
                "kind": s["kind"],
                "headers": s.get("headers", []),
                "rowCount": s.get("rowCount", 0),
                "secretPatternHits": s.get("secretPatternHits", 0),
            }
            for s in sheets
        ],
        "canonicalExpected": len(CANONICAL_QIDS),
        "matchedCount": len(matched),
        "matchedQuestionIds": matched,
        "missingQuestionIds": missing,
        "extraQuestionIds": extra,
        "answeredCanonicalCount": len(answered_qids),
        "answeredQuestionIds": answered_qids,
        "unansweredMatched": sorted(set(matched) - set(answered_qids)),
        "statusCounts": dict(status_counts),
        "rowsWithSource": with_source,
        "rowsWithStatus": with_status,
        "qaRowCount": len(qa_rows),
        "fieldsStrong": sorted(strong),
        "fieldsPartial": sorted(partial),
        "fieldsUnsupported": sorted(unsupported),
        "sectionHits": dict(section_hits),
        "gapEvidence": gap_evidence,
        "suggestedAdditionIds": suggested_ids,
        "secretPatternHitsTotal": sum(s.get("secretPatternHits", 0) for s in sheets),
    }


def main():
    root = Path(sys.argv[1] if len(sys.argv) > 1 else "resources/historical-rfps")
    excel_dir = root / "source" / "excel"
    results = []
    for path in sorted(excel_dir.glob("*.xlsx")):
        print(f"Auditing {path.name}…", file=sys.stderr)
        results.append(audit_workbook(path))

    strong_any: set[str] = set()
    partial_any: set[str] = set()
    for r in results:
        strong_any.update(r["fieldsStrong"])
        partial_any.update(r["fieldsPartial"])
    partial_any -= strong_any
    all_fields = set()
    for fs in Q_TO_FIELDS.values():
        all_fields.update(fs)

    section_rfp = Counter()
    for r in results:
        for s in r["sectionHits"]:
            section_rfp[s] += 1

    gap_recurrence = Counter()
    for r in results:
        for k, v in r["gapEvidence"].items():
            if v:
                gap_recurrence[k] += 1

    qa_header_sigs = []
    for r in results:
        for s in r["sheetsSummary"]:
            if s["kind"] == "qa":
                qa_header_sigs.append("|".join(s["headers"]))

    out = {
        "datasets": results,
        "combined": {
            "datasetCount": len(results),
            "fieldsStrongAny": sorted(strong_any),
            "fieldsPartialAny": sorted(partial_any),
            "fieldsUnsupportedAll": sorted(all_fields - strong_any - partial_any),
            "fieldCoverageStrong": len(strong_any),
            "fieldCoveragePartial": len(partial_any),
            "fieldCoverageTotalCanonical": len(all_fields),
            "sectionRfpCounts": dict(section_rfp),
            "gapRecurrenceAcrossDatasets": dict(gap_recurrence),
            "qaHeaderSignatures": sorted(set(qa_header_sigs)),
            "qaSchemaConsistent": len(set(qa_header_sigs)) == 1,
            "suggestedIdUnion": sorted({i for r in results for i in r["suggestedAdditionIds"]}),
        },
    }
    out_path = root / "derived" / "audit-raw.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"ok": True, "datasets": len(results), "path": str(out_path)}))


if __name__ == "__main__":
    main()
