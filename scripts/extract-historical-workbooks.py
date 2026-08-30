#!/usr/bin/env python3
"""
Extract historical RFP Excel workbooks to JSON for PostgreSQL import.
Read-only on source files. Does not touch ProjectFacts.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# Keep in sync with src/schema/questionBankSeed.ts
CANONICAL_QIDS = {
    "0.1", "0.2", "0.3", "0.4", "0.5", "0.6", "0.7",
    "1.1", "1.2", "1.3", "1.4", "1.5",
    "2.1", "2.2", "2.3",
    "3.1", "3.2", "3.3", "3.4", "3.5",
    "4.1", "4.2", "4.3", "4.4", "4.5",
    "5.1", "5.2", "5.3", "5.4", "5.5",
    "6.1", "6.2", "6.3", "6.4", "6.5", "6.6",
    "7.1", "7.2", "7.3",
    "8.1", "8.2", "8.3", "8.4",
    "9.1", "9.2", "9.3", "9.4",
    "10.1", "10.2", "10.3", "10.4",
    "11.1", "11.2", "11.3",
    "12.1", "12.2", "12.3", "12.4", "12.5", "12.6", "12.7", "12.8",
}

Q_TO_FIELDS: dict[str, list[str]] = {
    "0.1": ["documentType"], "0.2": ["documentTitle"], "0.3": ["beneficiaryEntity"],
    "0.4": ["tenderNumber"], "0.5": ["proposalDeadline"], "0.6": [], "0.7": ["referenceTemplateId"],
    "1.1": ["currentSituation"], "1.2": ["painPoints"], "1.3": ["businessNeedRationale"],
    "1.4": ["businessObjectives"], "1.5": ["previousPhases"],
    "2.1": ["engagementType"], "2.2": ["engagementPhases"], "2.3": ["engagementDuration"],
    "3.1": ["beneficiaryEntity"], "3.2": ["users"], "3.3": ["stakeholderRoles"],
    "3.4": ["approvers", "uatOwners"], "3.5": ["postGoLiveOwner"],
    "4.1": ["inScope"], "4.2": ["outOfScope"], "4.3": ["bidderResponsibilities"],
    "4.4": ["entityResponsibilities"], "4.5": ["assumptionsDependenciesConstraints"],
    "5.1": ["functionalModules"], "5.2": ["keyWorkflows"], "5.3": ["reportingNeeds"],
    "5.4": ["caseManagementNeeds"], "5.5": ["aiFeatures"],
    "6.1": ["hostingModel"], "6.2": ["integrations"], "6.3": ["securityRequirements"],
    "6.4": ["performanceAvailabilityTargets"], "6.5": ["dataMigrationNeeds"], "6.6": [],
    "7.1": ["deliverableItems"], "7.2": ["deliverableFormats"], "7.3": ["deliverableApprovers"],
    "8.1": ["engagementPhases", "engagementDuration"], "8.2": ["uatRounds"],
    "8.3": ["acceptanceCriteria"], "8.4": ["rollbackPlanNeeded"],
    "9.1": ["supportPeriodAndHours"], "9.2": ["slaTiers"], "9.3": ["supportOperatingModel"],
    "9.4": ["supportPenalties"],
    "10.1": ["evaluationWeights"], "10.2": ["evaluationRules"],
    "10.3": ["pricingModelAndCostBreakdown"], "10.4": ["optionalItemsAndTaxes"],
    "11.1": ["legalTerms"], "11.2": ["jvSubcontractingRules"], "11.3": ["requiredAnnexes"],
    "12.1": ["riskNotes"], "12.2": ["riskNotes"], "12.3": ["riskNotes"], "12.4": [],
    "12.5": ["optionalItemsAndTaxes"], "12.6": [], "12.7": [], "12.8": ["requiredAnnexes"],
}

EVAL_ELIGIBILITY = {
    "rfp-22-egovt-2026-reengineering-ofa": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": False,
        "fullDocumentPdf": True,
        "pageLevelProvenance": True,
    },
    "pq-15-egovt-2026-sanad-ai": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": True,
        "fullDocumentPdf": True,
        "pageLevelProvenance": True,
    },
    "rfp-ssc-bpr": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": False,
        "fullDocumentPdf": False,
        "pageLevelProvenance": False,
    },
    "rfp-nur-v2-lakehouse": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": False,
        "fullDocumentPdf": False,
        "pageLevelProvenance": False,
    },
    "rfp-itas-vol2b": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": False,
        "fullDocumentPdf": True,
        "pageLevelProvenance": True,
    },
    "rfp-connectivity-ofa": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": False,
        "fullDocumentPdf": False,
        "pageLevelProvenance": False,
    },
    "rfp-17-egovt-2026-performance-assessment": {
        "questionCoverage": True,
        "fieldCoverage": True,
        "extractionGolden": True,
        "stageTbc": False,
        "fullDocumentPdf": True,
        "pageLevelProvenance": True,
    },
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def find_header_row(rows: list) -> tuple[int | None, list[str] | None]:
    for i, r in enumerate(rows):
        vals = [cell_str(c) for c in (r or [])]
        norms = [v.lower() for v in vals]
        if any(n == "question id" for n in norms) and any("question" in n for n in norms):
            return i, vals
    return None, None


def map_columns(headers: list[str]) -> dict[str, int]:
    norms = [re.sub(r"\s+", " ", h.strip().lower()) for h in headers]
    out: dict[str, int] = {}
    for i, n in enumerate(norms):
        if n == "question id" or n == "qid":
            out["question_id"] = i
        elif n in {"exact rami question", "question", "question text", "suggested question"}:
            out["question_text"] = i
        elif "answer" in n:
            out["answer"] = i
        elif n == "status":
            out["status"] = i
        elif n.startswith("source"):
            out["source"] = i
        elif n == "section":
            out["section"] = i
    return out


def make_answer_id(rfp_id: str, sheet: str, qid: str) -> str:
    # Collision-safe across datasets for reused 13.x–17.x IDs
    return f"{rfp_id}::{sheet}::{qid}"


def parse_sheet(rows: list, sheet_name: str, kind: str, rfp_id: str, excel_rel: str, has_pdf: bool):
    hdr_i, headers = find_header_row(rows)
    if hdr_i is None or headers is None:
        return []
    cols = map_columns(headers)
    out = []
    for row_offset, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        vals = [cell_str(c) for c in r]
        qid = vals[cols["question_id"]] if "question_id" in cols else ""
        qid = re.sub(r"^q\s*", "", qid.strip(), flags=re.I)
        qid = re.sub(r"\s+", "", qid)
        qtext = vals[cols["question_text"]] if "question_text" in cols else ""
        if not qid and not qtext:
            continue
        ans = vals[cols["answer"]] if "answer" in cols else ""
        status = vals[cols["status"]] if "status" in cols else ("Answered" if ans else "")
        if kind == "suggested_additions" and not status:
            status = "Suggested"
        source = vals[cols["source"]] if "source" in cols else ""
        section = vals[cols["section"]] if "section" in cols else ""
        is_canonical = kind == "qa" and qid in CANONICAL_QIDS
        canonical = qid if is_canonical else None
        mapped = Q_TO_FIELDS.get(canonical, []) if canonical else []
        out.append({
            "answerId": make_answer_id(rfp_id, sheet_name, qid),
            "historicalRfpId": rfp_id,
            "sourceSheet": sheet_name,
            "sourceSheetKind": kind,
            "sourceRow": row_offset,
            "sourceQuestionId": qid,
            "canonicalQuestionId": canonical,
            "isCanonical": is_canonical,
            "questionSectionLabel": section or None,
            "exactQuestionText": qtext,
            "answerText": ans,
            "extractionStatus": status or "Unknown",
            "sourceLocator": source or None,
            "provenanceClass": "REFERENCE",
            "mappedFieldIds": mapped,
            "excelRelPath": excel_rel,
            "pdfAvailable": has_pdf and kind == "qa",
        })
    return out


def extract_workbook(path: Path, rfp_id: str, excel_rel: str, has_pdf: bool):
    wb = load_workbook(path, read_only=True, data_only=True)
    answers = []
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        lower = name.lower()
        if name == "Rami Q&A" or (lower.startswith("rami") and "q" in lower and "&" in name):
            kind = "qa"
        elif "suggested" in lower:
            kind = "suggested_additions"
        else:
            continue
        answers.extend(parse_sheet(rows, name, kind, rfp_id, excel_rel, has_pdf))
    wb.close()
    return answers


def main():
    root = Path(sys.argv[1] if len(sys.argv) > 1 else "resources/historical-rfps")
    manifest = json.loads((root / "manifest.json").read_text(encoding="utf-8"))
    documents = []
    answers = []
    errors = []

    for res in manifest["resources"]:
        rfp_id = res["id"]
        excel_rel = res["excel"]["path"]
        excel_path = root / excel_rel
        if not excel_path.is_file():
            errors.append(f"missing excel: {excel_rel}")
            continue
        actual = sha256(excel_path)
        expected = res["excel"]["sha256"]
        if actual != expected:
            errors.append(f"excel hash mismatch {rfp_id}: expected {expected} got {actual}")
            continue

        pdf_rel = res["pdf"]["path"] if res.get("pdf") else None
        pdf_sha = None
        has_pdf = False
        if pdf_rel:
            pdf_path = root / pdf_rel
            if not pdf_path.is_file():
                errors.append(f"missing pdf: {pdf_rel}")
            else:
                pdf_sha = sha256(pdf_path)
                if pdf_sha != res["pdf"]["sha256"]:
                    errors.append(f"pdf hash mismatch {rfp_id}")
                else:
                    has_pdf = True

        eligibility = EVAL_ELIGIBILITY.get(
            rfp_id,
            {
                "questionCoverage": True,
                "fieldCoverage": True,
                "extractionGolden": False,
                "stageTbc": False,
                "fullDocumentPdf": has_pdf,
                "pageLevelProvenance": has_pdf,
            },
        )
        if not has_pdf:
            eligibility = {
                **eligibility,
                "fullDocumentPdf": False,
                "pageLevelProvenance": False,
            }

        documents.append({
            "historicalRfpId": rfp_id,
            "title": res["name"],
            "sourceType": res["sourceType"],
            "documentKinds": res.get("documentKinds", []),
            "intendedUse": res.get("intendedUse", []),
            "excelRelPath": excel_rel,
            "excelSha256": actual,
            "pdfRelPath": pdf_rel if has_pdf else None,
            "pdfSha256": pdf_sha if has_pdf else None,
            "hasPdf": has_pdf,
            "evaluationEligibility": eligibility,
            "manifestJson": res,
            "notes": res.get("notes", []) + ([] if has_pdf else ["PDF unavailable"]),
        })

        try:
            answers.extend(extract_workbook(excel_path, rfp_id, excel_rel, has_pdf))
        except Exception as e:
            errors.append(f"parse failed {rfp_id}: {e}")

    payload = {
        "ok": len(errors) == 0,
        "errors": errors,
        "documents": documents,
        "answers": answers,
        "counts": {
            "documents": len(documents),
            "answers": len(answers),
            "canonicalAnswers": sum(1 for a in answers if a["isCanonical"]),
            "noncanonicalAnswers": sum(1 for a in answers if not a["isCanonical"]),
        },
    }
    out = root / "derived" / "import-payload.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"ok": payload["ok"], "path": str(out), **payload["counts"], "errors": errors}))


if __name__ == "__main__":
    main()
